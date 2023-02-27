from django.shortcuts import render
from django.http import JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from .models import Product, ProductVariation
from datetime import datetime, timedelta
import pandas as pd
import random
from openpyxl import load_workbook


def index(request):
    result = []

    # Retrieve all products from the database
    allProducts = Product.objects.all()

    # Create a paginator with 3 items per page and get the requested page number
    paginator = Paginator(allProducts, 3) 
    page_number = request.GET.get('page') 
    page_obj = paginator.get_page(page_number)


    # Loop through each product on the requested page and retrieve its variations
    for product in page_obj:

        variation = Product.objects.get(name=product.name).variations.all()

        # Add the product and its variations to a dictionary and append it to the result list
        product_variance = {
            "item": product,
            "variation": variation,
        }
        result.append(product_variance)

    # Create a context dictionary with the result list and the page object
    context = {
        "data": result,
        'page_obj': page_obj,
    }
    
    # Render the homepage template with the context dictionary
    return render(request, 'excelperser/homePage.html', context=context)


@csrf_exempt
def addProduct(request):

    # Check if the request method is POST and if a file was uploaded
    if request.method == 'POST' and request.FILES.get('file'):

        file = request.FILES['file']

        # Check if the file type and size are valid
        if file.content_type not in ['application/vnd.ms-excel', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet']:
            return JsonResponse({'error': 'Invalid file type'})
        if file.size > 2 * 1024 * 1024:
            return JsonResponse({'error': 'File size exceeded'})
        


        try:

            # Load the workbook from the uploaded file and retrieve the first worksheet
            productFile = load_workbook(filename=file)
            sheet_name = productFile.sheetnames[0]
            dataInFile = productFile[sheet_name]

            # Loop through each row in the worksheet and add the products and variations to the database
            ok = False

            for line in dataInFile:

                # Skip the first row (header row)
                if not ok:
                    ok = True
                    continue

                # Get the product name, variation text, and stock from the current row
                name, variation_text, stock = line[0].value, line[1].value, line[2].value 


                # Check if the product already exists in the database
                products = Product.objects.filter(name=name)

                if products:
                    # If the product exists, check if the variation already exists
                    variation = ProductVariation.objects.filter(product_id = products[0].id, variation_text=variation_text)

                    if variation:
                        # If the variation already exists, update its stock
                        variation[0].stock += stock

                        utc_now = datetime.utcnow()      
                        india_offset = timedelta(hours=5, minutes=30)                  
                        variation[0].last_updated = utc_now + india_offset
                        products[0].last_updated = (utc_now + india_offset)

                        products[0].save()
                        variation[0].save()
                    else:
                        # If the variation doesn't exist, create a new one
                        newVariation = ProductVariation(product_id=products[0].id, variation_text=variation_text, stock=stock)
                        newVariation.save()

                else:
                    
                    # If the product doesn't exist, create a new one and a new variation

                    newProduct = Product(name=name, lowest_price=random.randint(10000, 99999))
                    newProduct.save()

                    newVariation = ProductVariation(product_id=newProduct.id, variation_text=variation_text, stock=stock)
                    newVariation.save()
         
            return JsonResponse({'success': True})
        
        except Exception as e:
            return JsonResponse({'error': str(e)})
            
    else:
        return JsonResponse({'error': 'Invalid request'})
